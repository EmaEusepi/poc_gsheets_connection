from flask import Flask, request, jsonify
from flask_cors import CORS
import operator
import math
import fnmatch
import tempfile
import os
import time

# Dipendenze opzionali per /eval_sheet
# pip install formulas openpyxl numpy
try:
    import openpyxl
    import formulas as formulas_lib
    import numpy as np
    EVAL_SHEET_AVAILABLE = True
except ImportError:
    EVAL_SHEET_AVAILABLE = False

app = Flask(__name__)
CORS(app)  # Necessario per chiamate da Google Sheets

# Dizionario delle operazioni supportate
OPERATIONS = {
    # Operazioni matematiche base
    'plus': lambda *args: sum(args),
    'minus': lambda a, b: a - b,
    'multiply': lambda *args: math.prod(args),
    'divide': lambda a, b: a / b if b != 0 else '#DIV/0!',
    'power': lambda a, b: a ** b,
    'mod': lambda a, b: a % b,
    
    # Operazioni di confronto
    'equals': lambda a, b: a == b,
    'greater': lambda a, b: a > b,
    'less': lambda a, b: a < b,
    'greater_equal': lambda a, b: a >= b,
    'less_equal': lambda a, b: a <= b,
    
    # Operazioni logiche
    'and': lambda *args: all(args),
    'or': lambda *args: any(args),
    'not': lambda a: not a,
    
    # Operazioni condizionali
    'if': lambda condition, true_val, false_val: true_val if condition else false_val,
    'iferror': lambda value, fallback=0: fallback if (isinstance(value, str) and str(value).startswith('#')) else value,
    
    # Funzioni matematiche
    'sqrt': lambda a: math.sqrt(a),
    'abs': lambda a: abs(a),
    'round': lambda a, decimals=0: round(a, int(decimals)),
    'floor': lambda a: math.floor(a),
    'ceil': lambda a: math.ceil(a),
    
    # Funzioni aggregate
    'max': lambda *args: max(args),
    'min': lambda *args: min(args),
    'average': lambda *args: sum(args) / len(args) if args else 0,
    'count': lambda *args: len(args),
    
    # Funzioni stringa
    'concat': lambda *args: ''.join(str(x) for x in args),
    'upper': lambda s: str(s).upper(),
    'lower': lambda s: str(s).lower(),
    'trim': lambda s: str(s).strip(),
    'len': lambda s: len(str(s)),
}

def match_criteria(value, criteria):
    """Confronta un valore con un criterio stile SUMIFS/COUNTIFS"""
    if criteria is None:
        return value is None

    criteria_str = str(criteria).strip()

    # Operatori di confronto
    if criteria_str.startswith('>='):
        try:
            return float(value) >= float(criteria_str[2:])
        except (ValueError, TypeError):
            return False
    elif criteria_str.startswith('<='):
        try:
            return float(value) <= float(criteria_str[2:])
        except (ValueError, TypeError):
            return False
    elif criteria_str.startswith('<>'):
        return str(value).lower() != criteria_str[2:].lower()
    elif criteria_str.startswith('>'):
        try:
            return float(value) > float(criteria_str[1:])
        except (ValueError, TypeError):
            return False
    elif criteria_str.startswith('<'):
        try:
            return float(value) < float(criteria_str[1:])
        except (ValueError, TypeError):
            return False

    # Wildcard matching (*, ?)
    if '*' in criteria_str or '?' in criteria_str:
        return fnmatch.fnmatch(str(value).lower(), criteria_str.lower())

    # Match esatto (case-insensitive per stringhe)
    if isinstance(value, str) and isinstance(criteria, str):
        return value.strip().lower() == criteria.strip().lower()

    # Confronto numerico
    try:
        return float(value) == float(criteria)
    except (ValueError, TypeError):
        return str(value).strip().lower() == criteria_str.lower()


def calc_sumifs(sum_range, criteria_pairs):
    """Implementazione di SUMIFS: somma condizionale con criteri multipli"""
    total = 0
    for i in range(len(sum_range)):
        match = True
        for pair in criteria_pairs:
            crit_range = pair['range']
            criteria = pair['criteria']
            if i >= len(crit_range):
                match = False
                break
            if not match_criteria(crit_range[i], criteria):
                match = False
                break
        if match:
            val = sum_range[i]
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
    return total


def parse_value(value):
    """Converte il valore nel tipo appropriato"""
    if value is None or value == '':
        return None
    
    # Prova a convertire in numero
    if isinstance(value, (int, float, bool)):
        return value
    
    value_str = str(value).strip()
    
    # Booleani
    if value_str.lower() == 'true':
        return True
    if value_str.lower() == 'false':
        return False
    
    # Numeri
    try:
        if '.' in value_str:
            return float(value_str)
        return int(value_str)
    except ValueError:
        return value_str

@app.route('/calc', methods=['POST', 'GET'])
def calculate():
    try:
        # Supporta sia GET che POST
        if request.method == 'POST':
            data = request.get_json()
        else:
            data = {
                'operation': request.args.get('operation'),
                'args': request.args.getlist('args')
            }
        
        operation = data.get('operation', '').lower()
        args_raw = data.get('args', [])
        
        # Valida l'operazione
        if not operation:
            return jsonify({'error': 'Operation is required'}), 400
        
        # SUMIFS: gestione speciale con payload strutturato
        if operation == 'sumifs':
            sum_range_raw = data.get('sum_range', [])
            criteria_pairs_raw = data.get('criteria_pairs', [])

            if not sum_range_raw or not criteria_pairs_raw:
                return jsonify({
                    'error': 'sumifs richiede sum_range e criteria_pairs nel payload'
                }), 400

            sum_range = [parse_value(v) for v in sum_range_raw]
            criteria_pairs = []
            for pair in criteria_pairs_raw:
                criteria_pairs.append({
                    'range': [parse_value(v) for v in pair['range']],
                    'criteria': parse_value(pair['criteria'])
                })

            result = calc_sumifs(sum_range, criteria_pairs)
            return jsonify({
                'result': result,
                'operation': operation
            })

        if operation not in OPERATIONS:
            return jsonify({
                'error': f'Unknown operation: {operation}',
                'available': list(OPERATIONS.keys())
            }), 400

        # Parse degli argomenti
        args = [parse_value(arg) for arg in args_raw]

        # Esegui l'operazione
        result = OPERATIONS[operation](*args)
        
        return jsonify({
            'result': result,
            'operation': operation,
            'args': args
        })
    
    except TypeError as e:
        return jsonify({
            'error': f'Invalid number of arguments: {str(e)}',
            'operation': operation
        }), 400
    except Exception as e:
        return jsonify({
            'error': str(e),
            'operation': operation
        }), 500

@app.route('/eval_sheet', methods=['POST'])
def eval_sheet():
    """Valuta un intero foglio: riceve formule + valori, restituisce risultati."""
    if not EVAL_SHEET_AVAILABLE:
        return jsonify({
            'error': 'Dipendenze mancanti. Installa con: pip install formulas openpyxl numpy'
        }), 501

    tmp_path = None
    try:
        start = time.time()
        data = request.get_json()

        formulas_grid = data.get('formulas', [])
        values_grid = data.get('values', [])

        if not values_grid:
            return jsonify({'error': 'values grid is required'}), 400

        num_rows = len(values_grid)
        num_cols = max(len(row) for row in values_grid) if values_grid else 0

        if num_rows == 0 or num_cols == 0:
            return jsonify({'error': 'Empty sheet'}), 400

        # Crea workbook temporaneo con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Model'

        formula_count = 0

        for r in range(num_rows):
            for c in range(num_cols):
                cell = ws.cell(row=r + 1, column=c + 1)

                # Controlla se c'e' una formula
                formula = ''
                if r < len(formulas_grid) and c < len(formulas_grid[r]):
                    formula = formulas_grid[r][c]

                if formula:
                    # openpyxl accetta formule senza '=' iniziale
                    cell.value = formula
                    formula_count += 1
                else:
                    # Valore diretto
                    val = values_grid[r][c] if c < len(values_grid[r]) else None
                    cell.value = parse_value(val)

        # Salva su file temporaneo
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        wb.save(tmp_path)

        # Calcola con la libreria formulas
        xl_model = formulas_lib.ExcelModel().loads(tmp_path).finish()
        solution = xl_model.calculate()

        # Leggi risultati dalla soluzione
        results = []
        for r in range(num_rows):
            row = []
            for c in range(num_cols):
                # formulas usa chiavi tipo "'[book]Model'!A1"
                # Ricava il nome colonna Excel (A, B, ..., Z, AA, AB, ...)
                col_letter = openpyxl.utils.get_column_letter(c + 1)
                cell_ref = "'[{}]Model'!{}{}".format(
                    os.path.basename(tmp_path), col_letter, r + 1
                )

                if cell_ref in solution:
                    val = solution[cell_ref]
                    # Converti tipi numpy in tipi Python nativi
                    if isinstance(val, np.ndarray):
                        val = val.item() if val.size == 1 else val.tolist()
                    if isinstance(val, (np.integer,)):
                        val = int(val)
                    elif isinstance(val, (np.floating,)):
                        val = float(val)
                    elif isinstance(val, (np.bool_,)):
                        val = bool(val)
                    row.append(val)
                else:
                    # Cella senza formula: usa il valore originale
                    val = values_grid[r][c] if c < len(values_grid[r]) else None
                    row.append(parse_value(val) if val is not None else '')
            results.append(row)

        elapsed_ms = int((time.time() - start) * 1000)

        return jsonify({
            'results': results,
            'stats': {
                'total_cells': num_rows * num_cols,
                'formula_cells': formula_count,
                'eval_time_ms': elapsed_ms
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        # Pulizia file temporaneo
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@app.route('/operations', methods=['GET'])
def list_operations():
    """Elenca tutte le operazioni disponibili"""
    ops = list(OPERATIONS.keys()) + ['sumifs']
    return jsonify({
        'operations': sorted(ops)
    })

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
