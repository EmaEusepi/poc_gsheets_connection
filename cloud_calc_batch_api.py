"""
Cloud Calc Batch API - Valutazione fogli interi
================================================
Server che riceve un intero foglio (formule congelate + valori)
dal Google Apps Script (evaluateSheet), risolve le dipendenze
e restituisce la griglia dei risultati calcolati.

Dipendenze:
    pip install flask flask-cors openpyxl formulas numpy

Avvio:  python cloud_calc_batch_api.py

Endpoint:
    POST /eval_sheet  - valuta un intero foglio
    GET  /health      - health check
    GET  /operations  - lista operazioni disponibili
"""

from __future__ import annotations

from flask import Flask, request, jsonify
from flask_cors import CORS
import tempfile
import os
import time
import re

import openpyxl
import formulas as formulas_lib
import numpy as np

app = Flask(__name__)
CORS(app)

# ---------------------------------------------------------------------------
# Mappa nomi funzione italiani -> inglesi (Google Sheets / Excel italiano)
# ---------------------------------------------------------------------------
IT_TO_EN_FUNCTIONS = {
    # Logiche / Condizionali
    'SE': 'IF',
    'SE.ERRORE': 'IFERROR',
    'SE.NON.DISP': 'IFNA',
    'E': 'AND',
    'O': 'OR',
    'NON': 'NOT',
    'SWITCH': 'SWITCH',
    'SCEGLI': 'CHOOSE',
    # Matematiche
    'SOMMA': 'SUM',
    'PRODOTTO': 'PRODUCT',
    'QUOZIENTE': 'QUOTIENT',
    'RESTO': 'MOD',
    'POTENZA': 'POWER',
    'RADQ': 'SQRT',
    'ASS': 'ABS',
    'ARROTONDA': 'ROUND',
    'ARROTONDA.PER.DIF': 'ROUNDDOWN',
    'ARROTONDA.PER.ECC': 'ROUNDUP',
    'INT': 'INT',
    'CASUALE': 'RAND',
    'CASUALE.TRA': 'RANDBETWEEN',
    'LOG': 'LOG',
    'LOG10': 'LOG10',
    'LN': 'LN',
    'EXP': 'EXP',
    'PI.GRECO': 'PI',
    # Aggregate
    'MAX': 'MAX',
    'MIN': 'MIN',
    'MEDIA': 'AVERAGE',
    'MEDIA.SE': 'AVERAGEIF',
    'MEDIA.PIU.SE': 'AVERAGEIFS',
    'CONTA.NUMERI': 'COUNT',
    'CONTA.VALORI': 'COUNTA',
    'CONTA.VUOTE': 'COUNTBLANK',
    'CONTA.SE': 'COUNTIF',
    'CONTA.PIU.SE': 'COUNTIFS',
    'SOMMA.SE': 'SUMIF',
    'SOMMA.PIU.SE': 'SUMIFS',
    'GRANDE': 'LARGE',
    'PICCOLO': 'SMALL',
    # Ricerca
    'CERCA.VERT': 'VLOOKUP',
    'CERCA.ORIZZ': 'HLOOKUP',
    'INDICE': 'INDEX',
    'CONFRONTA': 'MATCH',
    'RIF.INDIRETTO': 'INDIRECT',
    'SCARTO': 'OFFSET',
    # Testo
    'CONCATENA': 'CONCATENATE',
    'CONCAT': 'CONCAT',
    'UNISCI.STRINGA': 'TEXTJOIN',
    'SINISTRA': 'LEFT',
    'DESTRA': 'RIGHT',
    'STRINGA.ESTRAI': 'MID',
    'LUNGHEZZA': 'LEN',
    'MAIUSC': 'UPPER',
    'MINUSC': 'LOWER',
    'MAIUSC.INIZ': 'PROPER',
    'ANNULLA.SPAZI': 'TRIM',
    'SOSTITUISCI': 'SUBSTITUTE',
    'RIMPIAZZA': 'REPLACE',
    'TROVA': 'FIND',
    'RICERCA': 'SEARCH',
    'TESTO': 'TEXT',
    'VALORE': 'VALUE',
    # Data
    'OGGI': 'TODAY',
    'ADESSO': 'NOW',
    'ANNO': 'YEAR',
    'MESE': 'MONTH',
    'GIORNO': 'DAY',
    'ORA': 'HOUR',
    'MINUTO': 'MINUTE',
    'SECONDO': 'SECOND',
    'DATA': 'DATE',
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_value(value):
    """Converte il valore nel tipo appropriato."""
    if value is None or value == '':
        return None
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value).strip()
    if s.lower() == 'true':
        return True
    if s.lower() == 'false':
        return False
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return s


def translate_formula_it_to_en(formula):
    """Traduce una formula dalla sintassi italiana a quella inglese.
    - Sostituisce nomi funzione IT -> EN
    - Sostituisce ; con , come separatore argomenti (fuori dalle stringhe)
    """
    if not formula or not formula.startswith('='):
        return formula

    result = formula

    # Sostituisci nomi funzione (ordine decrescente per lunghezza
    # per evitare sostituzioni parziali, es: SOMMA.PIU.SE prima di SOMMA)
    sorted_funcs = sorted(IT_TO_EN_FUNCTIONS.keys(), key=len, reverse=True)
    for it_name in sorted_funcs:
        en_name = IT_TO_EN_FUNCTIONS[it_name]
        pattern = re.compile(re.escape(it_name) + r'\s*\(', re.IGNORECASE)
        result = pattern.sub(en_name + '(', result)

    # Sostituisci ; con , ma solo fuori dalle stringhe
    translated = []
    in_string = False
    for ch in result:
        if in_string:
            translated.append(ch)
            if ch == '"':
                in_string = False
        else:
            if ch == '"':
                in_string = True
                translated.append(ch)
            elif ch == ';':
                translated.append(',')
            else:
                translated.append(ch)

    return ''.join(translated)


def convert_formulas_value(val):
    """Converte i tipi della libreria formulas in tipi Python nativi serializzabili JSON."""
    if val is None:
        return ''

    # Ranges object della libreria formulas -> prendi il primo valore
    if hasattr(val, 'value'):
        val = val.value

    # numpy array
    if isinstance(val, np.ndarray):
        val = val.item() if val.size == 1 else val.tolist()

    # Tipi numpy scalari
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        f = float(val)
        return int(f) if f == int(f) else f
    if isinstance(val, (np.bool_,)):
        return bool(val)
    if isinstance(val, (np.str_,)):
        return str(val)

    # Booleani Python nativi (prima di int, perche' bool e' subclass di int)
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float, str)):
        return val

    # Fallback
    return str(val)


# ---------------------------------------------------------------------------
# Request timing
# ---------------------------------------------------------------------------

@app.before_request
def _start_timer():
    request._start_time = time.time()


@app.after_request
def _log_request_time(response):
    elapsed = time.time() - getattr(request, '_start_time', time.time())
    print(f"[{request.method} {request.path}] {response.status_code} - {elapsed*1000:.0f}ms")
    return response


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.route('/eval_sheet', methods=['POST'])
def eval_sheet():
    """Valuta un intero foglio: riceve formule + valori, restituisce risultati.

    Payload atteso:
    {
        "formulas": [["=SUM(A1:A2)", "", ...], ...],
        "values":   [[null, 42, "hello", ...], ...]
    }

    - formulas[r][c]: stringa formula (es "=SUM(A1:A2)") o "" se non e' formula
    - values[r][c]:   valore letterale della cella (usato dove formulas e' "")

    Risposta:
    {
        "results": [[...], ...],
        "stats": {"total_cells": N, "formula_cells": N, "eval_time_ms": N}
    }
    """
    tmp_path = None
    try:
        start = time.time()
        data = request.get_json()

        formulas_grid = data.get('formulas', [])
        values_grid = data.get('values', [])

        if not values_grid:
            return jsonify({'error': 'values grid is required'}), 400

        num_rows = len(values_grid)
        num_cols = max(len(row) for row in values_grid) if values_grid else 0

        if num_rows == 0 or num_cols == 0:
            return jsonify({'error': 'Empty sheet'}), 400

        # ----- Costruisci workbook temporaneo con openpyxl -----
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Model'

        formula_count = 0

        for r in range(num_rows):
            for c in range(num_cols):
                cell = ws.cell(row=r + 1, column=c + 1)

                # Controlla se c'e' una formula
                formula = ''
                if r < len(formulas_grid) and c < len(formulas_grid[r]):
                    formula = formulas_grid[r][c]

                if formula:
                    cell.value = translate_formula_it_to_en(formula)
                    formula_count += 1
                else:
                    val = values_grid[r][c] if c < len(values_grid[r]) else None
                    cell.value = parse_value(val)

        # ----- Salva su file temporaneo e calcola -----
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        wb.save(tmp_path)

        xl_model = formulas_lib.ExcelModel().loads(tmp_path).finish()
        solution = xl_model.calculate()

        # ----- Costruisci lookup normalizzato dalla solution -----
        # La libreria formulas usa chiavi tipo "'[book.xlsx]Sheet'!A1"
        cell_pattern = re.compile(r"!([A-Z]+\d+)$", re.IGNORECASE)
        sheet_pattern = re.compile(r"\](.+?)'!", re.IGNORECASE)

        solution_map = {}
        for key, val in solution.items():
            cell_match = cell_pattern.search(str(key))
            sheet_match = sheet_pattern.search(str(key))
            if cell_match:
                cell_name = cell_match.group(1).upper()
                sheet_name = sheet_match.group(1).upper() if sheet_match else 'MODEL'
                normalized_key = f"{sheet_name}!{cell_name}"
                solution_map[normalized_key] = val

        # ----- Leggi risultati -----
        results = []
        for r in range(num_rows):
            row = []
            for c in range(num_cols):
                col_letter = openpyxl.utils.get_column_letter(c + 1)
                lookup_key = f"MODEL!{col_letter}{r + 1}"

                if lookup_key in solution_map:
                    val = solution_map[lookup_key]
                    val = convert_formulas_value(val)
                    row.append(val)
                else:
                    val = values_grid[r][c] if c < len(values_grid[r]) else None
                    row.append(parse_value(val) if val is not None else '')
            results.append(row)

        elapsed_ms = int((time.time() - start) * 1000)

        # Debug opzionale
        debug_info = {}
        if data.get('debug'):
            debug_info = {
                'raw_solution_keys': [str(k) for k in list(solution.keys())[:50]],
                'normalized_keys': list(solution_map.keys())[:50],
            }

        response_data = {
            'results': results,
            'stats': {
                'total_cells': num_rows * num_cols,
                'formula_cells': formula_count,
                'eval_time_ms': elapsed_ms
            }
        }
        if debug_info:
            response_data['debug'] = debug_info

        return jsonify(response_data)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'mode': 'batch_sheet'})


@app.route('/operations', methods=['GET'])
def list_operations():
    """Lista delle funzioni Excel supportate dalla libreria formulas."""
    return jsonify({
        'operations': sorted(IT_TO_EN_FUNCTIONS.keys()),
        'note': 'La libreria formulas supporta la maggior parte delle funzioni Excel standard. '
                'Le formule italiane vengono tradotte automaticamente in inglese.'
    })


if __name__ == '__main__':
    print("Cloud Calc Batch API - Valutazione fogli interi")
    print("Endpoints:")
    print("  POST /eval_sheet  - valuta un intero foglio")
    print("  GET  /health")
    print("  GET  /operations")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
